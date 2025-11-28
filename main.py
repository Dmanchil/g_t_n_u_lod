import camelot
import openpyxl 
import telebot
from datetime import datetime, timedelta 
import json
import os
import urllib.request

token='8269738099:AAETqsa8WwNzhBfVH2zLay7_svsH_DLQDTc'
#8347380655:AAE56FocrVCTzAY39vc4QOo9Oz0IsZttcBw ориг
#8269738099:AAETqsa8WwNzhBfVH2zLay7_svsH_DLQDTc тест
#8478375967:AAEA90do_76J-rG0tgRwKSuvD2uXXUk1JsY атпешки 23-1




bot=telebot.TeleBot(token)



#####################


#                                                                         ВАЖНО!!!
# 1 - группа Эг-24-1
# 2 - группа АТП-23-1

gr="1"#Какая именно группа 



#    "1":[0]
#    "группа" : [номер таблицы, колонка по горизонтали с группой]
group={"1":[1,5],"2":[2,9]}






@bot.message_handler(content_types=['text'])
def aaa(message):
    

    if(message.text == "/start"):
        keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
        button_support = telebot.types.KeyboardButton(text="Расписание")
        keyboard.add(button_support)
        bot.send_message(message.chat.id, "Привет", reply_markup=keyboard) 

    elif(message.text == "Расписание"):
        bot.send_message(1482529830,f"@{message.from_user.username}")

        kb1 = telebot.types.InlineKeyboardMarkup()#Клавиатуа


        #Добавляем кнопки


        b1= telebot.types.InlineKeyboardButton(text=f"<--",callback_data=f"<{datetime_full(-3)}")
        b2= telebot.types.InlineKeyboardButton(text=f"Сегодня",callback_data=f"{datetime_full(0)}")
        b3= telebot.types.InlineKeyboardButton(text=f"-->",callback_data=f">{datetime_full(3)}")

        b6= telebot.types.InlineKeyboardButton(text=f"{need_day(1)}",callback_data=f"{datetime_full(1)}")
        b7= telebot.types.InlineKeyboardButton(text=f"{need_day(2)}",callback_data=f"{datetime_full(2)}")
        b8= telebot.types.InlineKeyboardButton(text=f"{need_day(3)}",callback_data=f"{datetime_full(3)}")
        b9= telebot.types.InlineKeyboardButton(text=f"{need_day(-1)}",callback_data=f"{datetime_full(-1)}")
        b10= telebot.types.InlineKeyboardButton(text=f"{need_day(-2)}",callback_data=f"{datetime_full(-2)}")
        b11= telebot.types.InlineKeyboardButton(text=f"{need_day(-3)}",callback_data=f"{datetime_full(-3)}")
        
        #Вставляем в клавиатуру
        kb1.add(b6,b7,b8,b9,b10,b11,b1,b2,b3)
        bot.send_message(message.chat.id, "Привет", reply_markup=kb1) #Выводим клавиатуру

    

@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    if call.data[:1:] == f"<":
        
        time = datetime.strptime(call.data[1::],"%Y%m%d")

        kb1 = telebot.types.InlineKeyboardMarkup()#Клавиатуа


        b1= telebot.types.InlineKeyboardButton(text=f"<--",callback_data=f"<{plus_time_eng(time,-6)}")
        b2= telebot.types.InlineKeyboardButton(text=f"Сегодня",callback_data=f"{datetime_full(0)}")
        b3= telebot.types.InlineKeyboardButton(text=f"-->",callback_data=f">{plus_time_eng(time,-1)}")

        b6= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,-1)}",callback_data=f"{plus_time_eng(time,-1)}")
        b7= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,-2)}",callback_data=f"{plus_time_eng(time,-2)}")
        b8= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,-3)}",callback_data=f"{plus_time_eng(time,-3)}")
        b9= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,-4)}",callback_data=f"{plus_time_eng(time,-4)}")
        b10= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,-5)}",callback_data=f"{plus_time_eng(time,-5)}")
        b11= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,-6)}",callback_data=f"{plus_time_eng(time,-6)}")
        
        kb1.add(b6,b7,b8,b9,b10,b11,b1,b2,b3)

        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=f"Расписание.\n{datetime.now()}", reply_markup=kb1)
    elif call.data == f"20251129":
        bot.send_message(call.message.chat.id, "Создатели расписания ебланы, терпим их тупизм")
        return

    elif call.data[:1:] == f">":
        
        time = datetime.strptime(call.data[1::],"%Y%m%d")

        kb1 = telebot.types.InlineKeyboardMarkup()#Клавиатуа


        b1= telebot.types.InlineKeyboardButton(text=f"<--",callback_data=f"<{plus_time_eng(time,1)}")
        b2= telebot.types.InlineKeyboardButton(text=f"Сегодня",callback_data=f"{datetime_full(0)}")
        b3= telebot.types.InlineKeyboardButton(text=f"-->",callback_data=f">{plus_time_eng(time,6)}")

        b6= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,1)}",callback_data=f"{plus_time_eng(time,1)}")
        b7= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,2)}",callback_data=f"{plus_time_eng(time,2)}")
        b8= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,3)}",callback_data=f"{plus_time_eng(time,3)}")
        b9= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,4)}",callback_data=f"{plus_time_eng(time,4)}")
        b10= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,5)}",callback_data=f"{plus_time_eng(time,5)}")
        b11= telebot.types.InlineKeyboardButton(text=f"{plus_time(time,6)}",callback_data=f"{plus_time_eng(time,6)}")
        
        kb1.add(b6,b7,b8,b9,b10,b11,b1,b2,b3)

        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=f"Расписание.\n{datetime.now()}", reply_markup=kb1)

    else:
        send_mes(call)



    

def send_mes(call):
    url=f"https://gtnu.ru/wp-content/uploads/rasp/{call.data}.pdf"
    pdf=f'{call.data}.pdf'
    try:
        urllib.request.urlretrieve(url, pdf)
    except urllib.error.HTTPError:

        bot.send_message(call.from_user.id, f"Расписания нет. Отдыхай, разрешаю👍")
        return None

        
    

    desting = f'{call.data}.xlsx'

    a = camelot.read_pdf(pdf, pages='all')
    os.remove(pdf)

    if list(a) == []:

        bot.send_message(call.from_user.id, f"Расписания нет, либо создатели расписания ебланы")
        return      
    
    a[group[gr][0]].df.to_excel(desting)
    
    #открываем фаил exel
    fff=openpyxl.load_workbook(desting) 
    os.remove(desting) 
    f=fff.active

    #Ищем сколько всего пар
    g=[]
    for i in range(f.max_row-2):
        if f.cell(row=f.max_row-i,column=group[gr][1]).value is None:
            exit
        else:
            g.append([int(f.cell(row=f.max_row-i,column=group[gr][1]).value),[f.max_row-i,group[gr][1]]])
    
    g = g[::-1] 
    g.append([len(g)+1,[0,0]])
    #Состовляем сообщение
    mes=""

    for i in range(len(g)-1):
        #если нет пары

        if None is f.cell(row=g[i][1][0],column=g[i][1][1]+1).value:
            None
            #mes = mes + f"\n{g[i][0]}          {f.cell(row=g[i][1][0],column=2).value}\n        ———\n"





        #если препода 2

        
        elif len(f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()) == 3 and g[i][1][0]+2 != g[i+1][1][0] and f.cell(row=g[i][1][0]+2,column=g[i][1][1]+1).value != None:
            
            mes = mes + f"\n{g[i][0]}       {f.cell(row=g[i][1][0],column=2).value}\n    {f.cell(row=g[i][1][0],column=g[i][1][1]+1).value.splitlines()[0]}  —  {f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()[len(f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines())-1]}\n          {f.cell(row=g[i][1][0]+2,column=g[i][1][1]+1).value.splitlines()[1]}  —  {f.cell(row=g[i][1][0]+2,column=g[i][1][1]+1).value.splitlines()[0]}\n          {f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()[1]}  —  {f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()[0]}\n       {f.cell(row=g[i][1][0],column=g[i][1][1]+1).value.splitlines()[1]}\n"







                
                #если препода 2 но группы на разных парах

        elif len(f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()) == 3:
            
            mes = mes + f"\n{g[i][0]}       {f.cell(row=g[i][1][0],column=2).value}\n    {f.cell(row=g[i][1][0],column=g[i][1][1]+1).value.splitlines()[0]}  —  {f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()[len(f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines())-1]}\n          {f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()[1]}  —  {f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()[0]}\n       {f.cell(row=g[i][1][0],column=g[i][1][1]+1).value.splitlines()[1]}\n"




        #если ничего необычного
        else:#              1                                 08:30 - 09:55                                             Тех.мех.                                                        134                                                                                                                                                                 Лекции                                                                                                                                                                                      
            mes = mes + f"\n{g[i][0]}       {f.cell(row=g[i][1][0],column=2).value}\n    {f.cell(row=g[i][1][0],column=g[i][1][1]+1).value.splitlines()[0]}  —  {f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()[1]}\n          {f.cell(row=g[i][1][0]+1,column=g[i][1][1]+1).value.splitlines()[0]}\n       {f.cell(row=g[i][1][0],column=g[i][1][1]+1).value.splitlines()[len(f.cell(row=g[i][1][0],column=g[i][1][1]+1).value.splitlines())-1]}\n"
    

    bot.send_message(call.from_user.id, f"Расписание на  {call.data[6:8:]}.{call.data[4:6:]}\n" + mes)#Отправляем
    













######### Доп функции там всякие

def plus_time(time,n):
    return (time + timedelta(days=n)).strftime("%d.%m")

def plus_time_eng(time,n):
    return (time + timedelta(days=n)).strftime("%Y%m%d")

def need_day(n):
    now = datetime.now()  
    tomorrow = now + timedelta(days=n)
    return tomorrow.strftime("%d.%m")

def datetime_full(n):
    now = datetime.now()  
    tomorrow = now + timedelta(days=n)
    return tomorrow.strftime("%Y%m%d")


bot.polling()